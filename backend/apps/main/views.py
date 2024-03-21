# from formtools.wizard.views import SessionWizardView
from django.http import  HttpResponseRedirect
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from openpyxl import load_workbook
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger


from .models import *
from .forms import *

def home(request):
    return render(request, "users/home.html")

# =================================== The dashboard ===================================
@login_required
def dashboard(request):
    c_records = Child.objects.all()
    total_records = c_records.count()

    context = {
        "kids_registered": total_records,
        "total_no_of_kids": total_records,
    }
    return render(request, "main/dashboard.html", context)

# =================================== Fetch and display all children details ===================================
def child_list(request):
    queryset = Child.objects.all()

    # Search functionality
    search_query = request.GET.get('search')
    if search_query:
        queryset = queryset.filter(full_name__icontains=search_query)

    paginator = Paginator(queryset, 10)  # Show 10 records per page
    page = request.GET.get('page')

    try:
        c_records = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        c_records = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver last page of results.
        c_records = paginator.page(paginator.num_pages)

    return render(request, 'main/child/manage_child.html', {'c_records': c_records, 'table_title': 'Children MasterList'})

# =================================== Fetch and display selected child details ===================================
@login_required
def child_details(request, pk):
    record = Child.objects.get(pk=pk)
    age = record.calculate_age()
    
    context = {'table_title': 'Child Profile Report', 'record': record, 'age': age}
    return render(request, 'main/child/child_profile_rpt.html', context)

# =================================== Register Child ===================================
@login_required
@transaction.atomic
def register_child(request):
  if request.method == 'POST':
    form = ChildForm(request.POST, request.FILES)

    if form.is_valid(): 
      form.save()
      messages.info(request, "Record saved successfully!", extra_tags="bg-success")

    else:
      # Display form errors
      return render(request, 'main/child/child_frm.html', {'form': form})
  else:
    form = ChildForm()
  return render(request, 'main/child/child_frm.html', {'form_name':'Child Registration', 'form': form})

# =================================== Upload Profile Picture ===================================
def upload_profile_picture(request):
    # Retrieve the child object

    if request.method == 'POST':
        form = ChildProfilePictureForm(request.POST, request.FILES)
        if form.is_valid():
            # Create a new child profile picture
            child_picture = form.save(commit=False)
            child_picture.save()
            return redirect('child_list')
    else:
        form = ChildProfilePictureForm()

    context = {'form_name':'Upload Profile Pictures','form': form,}
    return render(request, 'main/child/profile_picture.html', context)

# =================================== Update Child data ===================================
@login_required
@transaction.atomic
def update_child(request, pk, template_name="main/child/child_frm.html"):
    try:
        c_record = Child.objects.get(pk=pk)
    except Child.DoesNotExist:
        messages.error(request, "Child record not found!", extra_tags="bg-danger")
        return redirect("child_list")  # Or a relevant error page

    # Update the form to handle pictures (assuming ChildForm)
    form = ChildForm(request.POST or None, request.FILES or None, instance=c_record)

    if form.is_valid():
        form.save()  # This will save the updated avatar if uploaded
        messages.info(request, "Record updated successfully!", extra_tags="bg-success")
        return redirect("child_list")  # Replace with the appropriate redirect URL
    
    context = {'form_name':'Child Registration', 'form': form, 'child': c_record}  # Include child record for display (optional)
    return render(request, template_name, context)

# =================================== Deleted selected child ===================================
@login_required
@transaction.atomic
def delete_child(request, pk):
    c_records = Child.objects.get(id=pk)
    c_records.delete()
    messages.info(request, "Record deleted successfully!", extra_tags="bg-danger")
    return HttpResponseRedirect(reverse("child_list"))

# =================================== Process and Import Excel data ===================================
@login_required
@transaction.atomic
# debug and refactor this code
def import_data(request):
    if request.method == 'POST':
        form = UploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            try:
                # Call process_and_import_data function
                process_and_import_data(excel_file)
                messages.success(request, "Data imported successfully!", extra_tags="bg-success")
            except Exception as e:
                messages.error(request, f"Error importing data: {e}", extra_tags="bg-danger")  # Handle unexpected errors
            return redirect("data_list")  # Replace with your redirect URL
    else:
        form = UploadForm()
    return render(request, 'main/child/import_frm.html', {'form_name': 'Import Excel Data', 'form': form})


# Function to import Excel data
def process_and_import_data(excel_file):
    try:
        wb = load_workbook(excel_file)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2):
            fname = row[0].value
            preferred_name = row[1].value
            residence = row[2].value
            tribe = row[3].value
            gender = row[4].value
            date_of_birth = row[5].value
            weight = row[6].value
            height = row[7].value
            avatar = row[8].value
            c_interest = row[9].value
            is_child_in_school = row[10].value
            name_of_the_school = row[11].value
            education_level = row[12].value
            child_class = row[13].value
            best_subject = row[14].value
            is_sponsored = row[15].value
            sponsorship_type = row[16].value
            father_name = row[17].value
            is_father_alive = row[18].value
            father_description = row[19].value
            mother_name = row[20].value
            is_mother_alive = row[21].value
            mother_description = row[22].value
            guardian = row[23].value
            guardian_contact = row[24].value
            relationship_with_guardian = row[25].value
            siblings = row[26].value
            background_info = row[27].value
            health_status = row[28].value
            responsibility = row[29].value
            relationship_with_christ = row[30].value
            religion = row[31].value
            prayer_request = row[32].value
            year_enrolled = row[33].value
            is_departed = row[34].value
            staff_comment = row[35].value
            compiled_by = row[36].value
            if fname is not None:
                obj = Child.objects.create(full_name=fname, preferred_name=preferred_name, residence=residence,
                                                 tribe=tribe, gender=gender, date_of_birth=date_of_birth, weight=weight,
                                                 height=height, avatar=avatar, c_interest=c_interest, is_child_in_school=is_child_in_school,
                                                 name_of_the_school=name_of_the_school, education_level=education_level,
                                                 child_class=child_class, best_subject=best_subject, is_sponsored=is_sponsored,
                                                 sponsorship_type=sponsorship_type, father_name=father_name, is_father_alive=is_father_alive,
                                                 father_description=father_description, mother_name=mother_name, is_mother_alive=is_mother_alive,
                                                 mother_description=mother_description, guardian=guardian, guardian_contact=guardian_contact, 
                                                 relationship_with_guardian=relationship_with_guardian, siblings=siblings, background_info=background_info,
                                                 health_status=health_status, responsibility=responsibility, relationship_with_christ=relationship_with_christ,
                                                 religion=religion, prayer_request= prayer_request, year_enrolled=year_enrolled, is_departed=is_departed, staff_comment=staff_comment,
                                                 compiled_by=compiled_by)
                obj.save()
    except Exception as e:
        raise e  # Reraise the exception for better error handling at the view level


# =================================== Fetch and display imported data ===================================
@login_required
def import_details(request):
    records = Child.objects.all()
    return render(request, 'main/child/data_list.html', {'table_title': 'Imported Excel Data', 'records': records})

# =================================== Delete selected individual ===================================
@login_required
@transaction.atomic
def delete_excel_data(request, pk):
    record_imported = Child.objects.get(id=pk)
    record_imported.delete()
    messages.info(request, "Record deleted!", extra_tags="bg-danger")
    return HttpResponseRedirect(reverse("data_list"))

# =================================== Delete all records at once ===================================
@login_required
@transaction.atomic
def delete_confirmation(request):
    if request.method == 'POST':
        Child.objects.all().delete()
        messages.info(request, "All records deleted!", extra_tags="bg-danger")
        return HttpResponseRedirect(reverse("data_list"))
    
# =================================== More ===================================
